# ===========================================
# routes/relatorios_routes.py - Relatorios Analiticos
# ===========================================

from datetime import datetime, timedelta
import os
import shutil
import csv
import io
import logging
from flask import Blueprint, jsonify, request, Response
from sqlalchemy import func
from repositories import relatorio_repository

relatorios_bp = Blueprint('relatorios', __name__, url_prefix='/api/relatorios')
logger = logging.getLogger(__name__)


def _excel_engine_disponivel():
    """Retorna engine Excel disponível para pandas (prioriza openpyxl)."""
    try:
        import openpyxl  # noqa: F401
        return 'openpyxl'
    except Exception:
        pass
    try:
        import xlsxwriter  # noqa: F401
        return 'xlsxwriter'
    except Exception:
        return None


def _aplicar_estilo_excel(writer, engine_excel, configuracoes_planilhas):
    """
    Aplica estilo profissional nas planilhas exportadas:
    - cabeçalho destacado
    - autofiltro
    - congelar linha de cabeçalho
    - largura automática de colunas
    - formatação de moeda/data
    """
    if engine_excel == 'xlsxwriter':
        workbook = writer.book
        header_fmt = workbook.add_format({
            'bold': True,
            'font_color': '#FFFFFF',
            'bg_color': '#0F4C5C',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        text_fmt = workbook.add_format({'border': 1, 'valign': 'vcenter'})
        money_fmt = workbook.add_format({'num_format': 'R$ #,##0.00', 'border': 1, 'valign': 'vcenter'})
        date_fmt = workbook.add_format({'num_format': 'dd/mm/yyyy', 'border': 1, 'valign': 'vcenter'})
        datetime_fmt = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm', 'border': 1, 'valign': 'vcenter'})
        integer_fmt = workbook.add_format({'num_format': '0', 'border': 1, 'valign': 'vcenter', 'align': 'center'})

        for conf in configuracoes_planilhas:
            nome = conf['sheet_name']
            df = conf['df']
            currency_cols = set(conf.get('currency_cols', []))
            date_cols = set(conf.get('date_cols', []))
            datetime_cols = set(conf.get('datetime_cols', []))
            integer_cols = set(conf.get('integer_cols', []))

            worksheet = writer.sheets.get(nome)
            if worksheet is None:
                continue

            max_row = len(df) + 1
            max_col = len(df.columns)
            if max_col <= 0:
                continue

            worksheet.freeze_panes(1, 0)
            worksheet.autofilter(0, 0, max_row - 1, max_col - 1)
            worksheet.set_row(0, 22, header_fmt)

            for col_idx, col_name in enumerate(df.columns):
                serie = df[col_name]
                tamanho_header = len(str(col_name))
                tamanho_conteudo = 0
                if not serie.empty:
                    tamanho_conteudo = int(serie.astype(str).map(len).max())
                largura = max(tamanho_header, tamanho_conteudo, 10) + 2
                largura = min(largura, 48)

                col_fmt = text_fmt
                if col_name in currency_cols:
                    col_fmt = money_fmt
                elif col_name in datetime_cols:
                    col_fmt = datetime_fmt
                elif col_name in date_cols:
                    col_fmt = date_fmt
                elif col_name in integer_cols:
                    col_fmt = integer_fmt

                worksheet.set_column(col_idx, col_idx, largura, col_fmt)
                worksheet.write(0, col_idx, col_name, header_fmt)

    elif engine_excel == 'openpyxl':
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='0F4C5C')
        header_align = Alignment(horizontal='center', vertical='center')
        body_align = Alignment(vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for conf in configuracoes_planilhas:
            nome = conf['sheet_name']
            df = conf['df']
            currency_cols = set(conf.get('currency_cols', []))
            date_cols = set(conf.get('date_cols', []))
            datetime_cols = set(conf.get('datetime_cols', []))
            integer_cols = set(conf.get('integer_cols', []))

            ws = writer.sheets.get(nome)
            if ws is None:
                continue

            ws.freeze_panes = 'A2'
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                serie = df[col_name]
                tamanho_header = len(str(col_name))
                tamanho_conteudo = 0
                if not serie.empty:
                    tamanho_conteudo = int(serie.astype(str).map(len).max())
                largura = max(tamanho_header, tamanho_conteudo, 10) + 2
                ws.column_dimensions[col_letter].width = min(largura, 48)

                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if col_name in integer_cols:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = body_align
                    if col_name in currency_cols:
                        cell.number_format = 'R$ #,##0.00'
                    elif col_name in datetime_cols:
                        cell.number_format = 'DD/MM/YYYY HH:mm'
                    elif col_name in date_cols:
                        cell.number_format = 'DD/MM/YYYY'


def _parse_data(valor, fim_do_dia=False):
    """Converte YYYY-MM-DD em datetime para filtro opcional."""
    if not valor:
        return None
    data = datetime.strptime(valor, '%Y-%m-%d')
    if fim_do_dia:
        return data + timedelta(days=1) - timedelta(microseconds=1)
    return data


def _parse_data_dia(valor):
    if not valor:
        return datetime.now()
    return datetime.strptime(valor, '%Y-%m-%d')


def _resolver_profissional_ativo(nome_informado):
    return relatorio_repository.resolver_profissional_ativo(nome_informado)


def _base_producao_query(data_inicio=None, data_fim=None):
    """Query base para reaproveitar filtro por periodo e identificacao do profissional."""
    return relatorio_repository.base_producao_query(data_inicio, data_fim)


def _periodo_inicio_fim(data_ref, tipo):
    return relatorio_repository.periodo_inicio_fim(data_ref, tipo)


def _detalhes_resumo_profissional(profissional_nome, data_inicio, data_fim, limite=100):
    return relatorio_repository.detalhes_resumo_profissional(profissional_nome, data_inicio, data_fim, limite=limite)


def _gerar_html_profissional(nome_profissional, cnpj, periodo_label, resumo, detalhes):
    """HTML do relatorio individual por profissional."""
    linhas = ""
    for item in detalhes:
        linhas += f"""
            <tr>
                <td>{item['ordem_id']}</td>
                <td>{item['data_referencia']}</td>
                <td>{item['cliente']}</td>
                <td>{item['descricao_servico']}</td>
                <td style="text-align:right;">R$ {item['valor_servico']:.2f}</td>
            </tr>
        """

    if not linhas:
        linhas = """
            <tr>
                <td colspan="5" style="text-align:center; color:#777;">Nenhum servico no periodo.</td>
            </tr>
        """

    return f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
      <meta charset="UTF-8">
      <title>Relatorio - {nome_profissional}</title>
      <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; color: #222; }}
        h1 {{ color: #0a3147; margin-bottom: 2px; }}
        h2 {{ color: #0a3147; margin-top: 24px; }}
        .meta {{ margin-bottom: 16px; color: #555; }}
        .card {{ background: #f7f7f7; border: 1px solid #ddd; border-radius: 8px; padding: 12px; margin: 16px 0; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #0a3147; color: #fff; padding: 8px; text-align: left; }}
        td {{ border-bottom: 1px solid #ddd; padding: 8px; }}
      </style>
    </head>
    <body>
      <h1>Relatorio de Producao por Profissional</h1>
      <div class="meta">
        <div><strong>Profissional:</strong> {nome_profissional}</div>
        <div><strong>CNPJ:</strong> {cnpj or 'Nao informado'}</div>
        <div><strong>Periodo:</strong> {periodo_label}</div>
        <div><strong>Emitido em:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
      </div>
      <div class="card">
        <div><strong>Quantidade de servicos:</strong> {resumo['quantidade_servicos']}</div>
        <div><strong>Valor total:</strong> R$ {resumo['valor_total']:.2f}</div>
        <div><strong>Media por servico:</strong> R$ {resumo['media_por_servico']:.2f}</div>
      </div>
      <h2>Servicos do Periodo</h2>
      <table>
        <thead>
          <tr>
            <th>OS</th>
            <th>Data</th>
            <th>Cliente</th>
            <th>Servico</th>
            <th style="text-align:right;">Valor</th>
          </tr>
        </thead>
        <tbody>{linhas}</tbody>
      </table>
    </body>
    </html>
    """


@relatorios_bp.route('/producao-profissionais', methods=['GET'])
def relatorio_producao_profissionais():
    """
    Retorna producao por profissional:
    - quantidade de servicos executados
    - valor total gerado
    - media por servico
    """
    try:
        data_inicio = _parse_data(request.args.get('data_inicio'))
        data_fim = _parse_data(request.args.get('data_fim'), fim_do_dia=True)
        from models import ItemServico

        query_base, profissional_expr, _ = _base_producao_query(data_inicio, data_fim)
        query = query_base.with_entities(
            profissional_expr.label('profissional'),
            func.count(ItemServico.id).label('quantidade_servicos'),
            func.coalesce(func.sum(ItemServico.valor_servico), 0.0).label('valor_total'),
            func.coalesce(func.avg(ItemServico.valor_servico), 0.0).label('media_por_servico')
        )

        resultados = (
            query
            .group_by(profissional_expr)
            .order_by(func.coalesce(func.sum(ItemServico.valor_servico), 0.0).desc())
            .all()
        )

        return jsonify([
            {
                'profissional': r.profissional,
                'quantidade_servicos': int(r.quantidade_servicos or 0),
                'valor_total': float(r.valor_total or 0),
                'media_por_servico': float(r.media_por_servico or 0)
            }
            for r in resultados
        ])

    except ValueError:
        return jsonify({'erro': 'Formato de data invalido. Use YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/profissionais', methods=['GET'])
def listar_profissionais_producao():
    """Lista profissionais cadastrados/ativos para seleção na tela de produção."""
    try:
        termo = (request.args.get('termo') or '').strip()
        rows = relatorio_repository.listar_profissionais_ativos(termo, limite=100)

        return jsonify([{
            'profissional': row.nome,
            'quantidade_servicos': 0
        } for row in rows])
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/resumo-profissional', methods=['GET'])
def resumo_periodos_profissional():
    """
    Retorna resumo diário, semanal e mensal de um profissional.
    """
    try:
        profissional = (request.args.get('profissional') or '').strip()
        if not profissional:
            return jsonify({'erro': 'Parametro profissional é obrigatório.'}), 400
        if relatorio_repository.total_profissionais_ativos() == 0:
            return jsonify({'erro': 'Não há profissionais cadastrados/ativos para gerar relatório.'}), 400

        data_ref = _parse_data_dia((request.args.get('data_ref') or '').strip())
        profissional_resolvido = _resolver_profissional_ativo(profissional)
        if not profissional_resolvido:
            return jsonify({'erro': 'Profissional não encontrado na lista de cadastrados/ativos.'}), 404
        nome_resolvido = profissional_resolvido.nome

        inicio_dia, fim_dia = _periodo_inicio_fim(data_ref, 'dia')
        inicio_semana, fim_semana = _periodo_inicio_fim(data_ref, 'semana')
        inicio_mes, fim_mes = _periodo_inicio_fim(data_ref, 'mes')

        diario = _detalhes_resumo_profissional(nome_resolvido, inicio_dia, fim_dia)
        semanal = _detalhes_resumo_profissional(nome_resolvido, inicio_semana, fim_semana)
        mensal = _detalhes_resumo_profissional(nome_resolvido, inicio_mes, fim_mes)

        return jsonify({
            'profissional': nome_resolvido,
            'data_referencia': data_ref.strftime('%Y-%m-%d'),
            'periodos': {
                'diario': {
                    'intervalo': {
                        'inicio': inicio_dia.strftime('%Y-%m-%d'),
                        'fim': fim_dia.strftime('%Y-%m-%d')
                    },
                    **diario
                },
                'semanal': {
                    'intervalo': {
                        'inicio': inicio_semana.strftime('%Y-%m-%d'),
                        'fim': fim_semana.strftime('%Y-%m-%d')
                    },
                    **semanal
                },
                'mensal': {
                    'intervalo': {
                        'inicio': inicio_mes.strftime('%Y-%m-%d'),
                        'fim': fim_mes.strftime('%Y-%m-%d')
                    },
                    **mensal
                }
            }
        })
    except ValueError:
        return jsonify({'erro': 'Formato de data invalido. Use YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/resumo-profissional-periodo', methods=['GET'])
def resumo_profissional_periodo_personalizado():
    """
    Retorna resumo e serviços de um profissional em um intervalo personalizado
    com no máximo 31 dias.
    """
    try:
        profissional = (request.args.get('profissional') or '').strip()
        data_inicio_str = (request.args.get('data_inicio') or '').strip()
        data_fim_str = (request.args.get('data_fim') or '').strip()

        if not profissional:
            return jsonify({'erro': 'Parametro profissional é obrigatório.'}), 400
        if relatorio_repository.total_profissionais_ativos() == 0:
            return jsonify({'erro': 'Não há profissionais cadastrados/ativos para gerar relatório.'}), 400
        if not data_inicio_str or not data_fim_str:
            return jsonify({'erro': 'Informe data_inicio e data_fim.'}), 400

        data_inicio = _parse_data(data_inicio_str)
        data_fim = _parse_data(data_fim_str, fim_do_dia=True)
        if not data_inicio or not data_fim:
            return jsonify({'erro': 'Datas inválidas.'}), 400
        if data_fim < data_inicio:
            return jsonify({'erro': 'data_fim deve ser maior ou igual à data_inicio.'}), 400

        dias_intervalo = (data_fim.date() - data_inicio.date()).days + 1
        if dias_intervalo > 31:
            return jsonify({'erro': 'O período máximo permitido é de 31 dias.'}), 400

        profissional_resolvido = _resolver_profissional_ativo(profissional)
        if not profissional_resolvido:
            return jsonify({'erro': 'Profissional não encontrado na lista de cadastrados/ativos.'}), 404
        nome_resolvido = profissional_resolvido.nome
        bloco = _detalhes_resumo_profissional(nome_resolvido, data_inicio, data_fim)

        return jsonify({
            'profissional': nome_resolvido,
            'periodo': {
                'inicio': data_inicio.strftime('%Y-%m-%d'),
                'fim': data_fim.strftime('%Y-%m-%d'),
                'dias': dias_intervalo
            },
            **bloco
        })
    except ValueError:
        return jsonify({'erro': 'Formato de data invalido. Use YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/painel-dia', methods=['GET'])
def painel_dia():
    """
    Resumo operacional diário para Home:
    - OS abertas
    - OS concluídas hoje
    - faturamento hoje
    - saídas hoje
    - saldo do dia
    """
    try:
        from models import Ordem, Saida

        data_ref = _parse_data_dia((request.args.get('data') or '').strip())
        data_inicio = data_ref.replace(hour=0, minute=0, second=0, microsecond=0)
        data_fim = data_ref.replace(hour=23, minute=59, second=59, microsecond=999999)

        ordens_abertas = (
            Ordem.query
            .filter(~Ordem.status.in_(['Concluído', 'Garantia']))
            .count()
        )

        concluidas_hoje = (
            Ordem.query
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .count()
        )

        faturamento_row = (
            Ordem.query
            .with_entities(func.coalesce(func.sum(Ordem.total_geral), 0.0))
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .first()
        )
        faturamento_hoje = float((faturamento_row[0] if faturamento_row else 0) or 0)

        saidas_row = (
            Saida.query
            .with_entities(func.coalesce(func.sum(Saida.valor), 0.0))
            .filter(Saida.data >= data_inicio.date(), Saida.data <= data_fim.date())
            .first()
        )
        saidas_hoje = float((saidas_row[0] if saidas_row else 0) or 0)

        return jsonify({
            'data': data_inicio.strftime('%Y-%m-%d'),
            'ordens_abertas': int(ordens_abertas or 0),
            'ordens_concluidas_hoje': int(concluidas_hoje or 0),
            'faturamento_hoje': faturamento_hoje,
            'saidas_hoje': saidas_hoje,
            'saldo_hoje': faturamento_hoje - saidas_hoje
        })
    except ValueError:
        return jsonify({'erro': 'Formato de data invalido. Use YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/detalhe', methods=['GET'])
def relatorio_producao_profissional_detalhe():
    """Retorna resumo e detalhamento dos servicos de um profissional no periodo."""
    try:
        from models import ItemServico, Ordem, Cliente

        profissional = (request.args.get('profissional') or '').strip()
        if not profissional:
            return jsonify({'erro': 'Parametro profissional e obrigatorio.'}), 400

        data_inicio = _parse_data(request.args.get('data_inicio'))
        data_fim = _parse_data(request.args.get('data_fim'), fim_do_dia=True)
        cnpj = (request.args.get('cnpj') or '').strip()

        query_base, profissional_expr, data_ref_expr = _base_producao_query(data_inicio, data_fim)
        query_base = query_base.filter(profissional_expr == profissional)

        resumo_row = query_base.with_entities(
            func.count(ItemServico.id),
            func.coalesce(func.sum(ItemServico.valor_servico), 0.0),
            func.coalesce(func.avg(ItemServico.valor_servico), 0.0)
        ).first()

        resumo = {
            'profissional': profissional,
            'cnpj': cnpj,
            'quantidade_servicos': int((resumo_row[0] if resumo_row else 0) or 0),
            'valor_total': float((resumo_row[1] if resumo_row else 0) or 0),
            'media_por_servico': float((resumo_row[2] if resumo_row else 0) or 0)
        }

        detalhes_rows = query_base.with_entities(
            Ordem.id.label('ordem_id'),
            data_ref_expr.label('data_referencia'),
            Cliente.nome_cliente.label('cliente'),
            ItemServico.descricao_servico,
            ItemServico.valor_servico
        ).join(Cliente, Ordem.cliente_id == Cliente.id).order_by(data_ref_expr.desc()).all()

        detalhes = [{
            'ordem_id': row.ordem_id,
            'data_referencia': row.data_referencia.strftime('%d/%m/%Y') if row.data_referencia else '---',
            'cliente': row.cliente or '---',
            'descricao_servico': row.descricao_servico or '---',
            'valor_servico': float(row.valor_servico or 0)
        } for row in detalhes_rows]

        periodo_label = (
            f"{request.args.get('data_inicio') or 'inicio'} ate {request.args.get('data_fim') or 'hoje'}"
        )

        return jsonify({
            'resumo': resumo,
            'detalhes': detalhes,
            'periodo': periodo_label
        })

    except ValueError:
        return jsonify({'erro': 'Formato de data invalido. Use YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/enviar-contador', methods=['POST'])
def enviar_relatorios_profissionais_contador():
    """Envia um e-mail separado para cada profissional no periodo informado."""
    try:
        dados = request.json or {}
        remetente = (dados.get('email_cliente') or '').strip()
        senha_app = (dados.get('senha_app') or '').strip()
        email_contador = (dados.get('email_contador') or '').strip()
        profissionais = dados.get('profissionais') or []

        if not remetente or not senha_app or not email_contador:
            return jsonify({'erro': 'email_cliente, senha_app e email_contador sao obrigatorios.'}), 400
        if not isinstance(profissionais, list) or len(profissionais) == 0:
            return jsonify({'erro': 'Informe ao menos um profissional para envio.'}), 400

        data_inicio = _parse_data(dados.get('data_inicio'))
        data_fim = _parse_data(dados.get('data_fim'), fim_do_dia=True)

        from infrastructure.email_service import enviar_relatorio_email

        periodo_label = f"{dados.get('data_inicio') or 'inicio'} ate {dados.get('data_fim') or 'hoje'}"
        resultados = []

        for item in profissionais:
            nome = (item.get('profissional') or '').strip()
            cnpj = (item.get('cnpj') or '').strip()
            if not nome:
                continue

            # Reaproveita a mesma logica da rota de detalhe para montar o conteudo por profissional.
            request_args = {
                'profissional': nome,
                'data_inicio': dados.get('data_inicio'),
                'data_fim': dados.get('data_fim'),
                'cnpj': cnpj
            }

            detalhe_resp = relatorio_producao_profissional_detalhe_interno(request_args)
            resumo = detalhe_resp['resumo']
            html = _gerar_html_profissional(
                nome_profissional=nome,
                cnpj=cnpj,
                periodo_label=periodo_label,
                resumo=resumo,
                detalhes=detalhe_resp['detalhes']
            )

            sucesso, msg = enviar_relatorio_email(
                remetente=remetente,
                senha=senha_app,
                destinatario=email_contador,
                periodo=f"Producao {nome} ({periodo_label})",
                html=html,
                formato='html'
            )

            resultados.append({
                'profissional': nome,
                'cnpj': cnpj,
                'enviado': bool(sucesso),
                'mensagem': msg
            })

        enviados = sum(1 for r in resultados if r['enviado'])
        return jsonify({
            'mensagem': f'{enviados} de {len(resultados)} relatorios enviados.',
            'resultados': resultados
        })
    except ValueError:
        return jsonify({'erro': 'Formato de data invalido. Use YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


def relatorio_producao_profissional_detalhe_interno(args_dict):
    """Versao interna do detalhe para reuso no envio por e-mail."""
    from models import ItemServico, Ordem, Cliente

    profissional = (args_dict.get('profissional') or '').strip()
    data_inicio = _parse_data(args_dict.get('data_inicio'))
    data_fim = _parse_data(args_dict.get('data_fim'), fim_do_dia=True)
    cnpj = (args_dict.get('cnpj') or '').strip()

    query_base, profissional_expr, data_ref_expr = _base_producao_query(data_inicio, data_fim)
    query_base = query_base.filter(profissional_expr == profissional)

    resumo_row = query_base.with_entities(
        func.count(ItemServico.id),
        func.coalesce(func.sum(ItemServico.valor_servico), 0.0),
        func.coalesce(func.avg(ItemServico.valor_servico), 0.0)
    ).first()

    resumo = {
        'profissional': profissional,
        'cnpj': cnpj,
        'quantidade_servicos': int((resumo_row[0] if resumo_row else 0) or 0),
        'valor_total': float((resumo_row[1] if resumo_row else 0) or 0),
        'media_por_servico': float((resumo_row[2] if resumo_row else 0) or 0)
    }

    detalhes_rows = query_base.with_entities(
        Ordem.id.label('ordem_id'),
        data_ref_expr.label('data_referencia'),
        Cliente.nome_cliente.label('cliente'),
        ItemServico.descricao_servico,
        ItemServico.valor_servico
    ).join(Cliente, Ordem.cliente_id == Cliente.id).order_by(data_ref_expr.desc()).all()

    detalhes = [{
        'ordem_id': row.ordem_id,
        'data_referencia': row.data_referencia.strftime('%d/%m/%Y') if row.data_referencia else '---',
        'cliente': row.cliente or '---',
        'descricao_servico': row.descricao_servico or '---',
        'valor_servico': float(row.valor_servico or 0)
    } for row in detalhes_rows]

    return {'resumo': resumo, 'detalhes': detalhes}


@relatorios_bp.route('/fechamento-dia', methods=['POST'])
def fechamento_dia():
    """
    Gera fechamento diário operacional:
    - ranking de faturamento por profissional no dia
    - total geral do dia
    - backup manual imediato do banco
    """
    try:
        from models import ItemServico, Ordem

        dados = request.json or {}
        data_ref_str = (dados.get('data') or '').strip()
        data_ref = _parse_data(data_ref_str) if data_ref_str else datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        data_inicio = data_ref.replace(hour=0, minute=0, second=0, microsecond=0)
        data_fim = data_ref.replace(hour=23, minute=59, second=59, microsecond=999999)

        # Reaproveita a query base para manter consistência com o relatório principal.
        query_base, profissional_expr, _ = _base_producao_query(data_inicio, data_fim)
        ranking_rows = (
            query_base
            .with_entities(
                profissional_expr.label('profissional'),
                func.count(ItemServico.id).label('quantidade_servicos'),
                func.coalesce(func.sum(ItemServico.valor_servico), 0.0).label('valor_total'),
                func.coalesce(func.avg(ItemServico.valor_servico), 0.0).label('media_por_servico')
            )
            .group_by(profissional_expr)
            .order_by(func.coalesce(func.sum(ItemServico.valor_servico), 0.0).desc())
            .all()
        )

        ranking = [{
            'profissional': row.profissional,
            'quantidade_servicos': int(row.quantidade_servicos or 0),
            'valor_total': float(row.valor_total or 0),
            'media_por_servico': float(row.media_por_servico or 0)
        } for row in ranking_rows]
        total_geral = float(sum(item['valor_total'] for item in ranking))

        pagamentos_rows = (
            Ordem.query
            .with_entities(
                func.coalesce(func.nullif(func.trim(Ordem.forma_pagamento), ''), 'Não informado').label('forma_pagamento'),
                func.coalesce(func.sum(Ordem.total_geral), 0.0).label('valor_total'),
                func.count(Ordem.id).label('quantidade')
            )
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .group_by('forma_pagamento')
            .order_by(func.coalesce(func.sum(Ordem.total_geral), 0.0).desc())
            .all()
        )
        pagamentos = [{
            'forma_pagamento': row.forma_pagamento,
            'valor_total': float(row.valor_total or 0),
            'quantidade': int(row.quantidade or 0)
        } for row in pagamentos_rows]

        # Backup manual no mesmo clique de fechamento para reduzir risco operacional.
        basedir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        db_path = os.path.join(basedir, 'database.db')
        backup_dir = os.path.join(basedir, 'backups')
        os.makedirs(backup_dir, exist_ok=True)

        backup_arquivo = None
        if os.path.exists(db_path):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_arquivo = f'fechamento_backup_{timestamp}.db'
            destino = os.path.join(backup_dir, backup_arquivo)
            shutil.copy2(db_path, destino)

        return jsonify({
            'data': data_inicio.strftime('%Y-%m-%d'),
            'total_geral': total_geral,
            'quantidade_profissionais': len(ranking),
            'ranking': ranking,
            'pagamentos': pagamentos,
            'backup_arquivo': backup_arquivo
        })

    except ValueError:
        return jsonify({'erro': 'Formato de data invalido. Use YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/exportar-csv', methods=['GET'])
def exportar_csv_producao_pagamentos():
    """
    Exporta CSV mensal para contador com:
    - ranking de produção por profissional
    - resumo por forma de pagamento
    """
    try:
        from models import ItemServico, Ordem

        mes = (request.args.get('mes') or '').strip()  # YYYY-MM
        if mes:
            data_base = datetime.strptime(f'{mes}-01', '%Y-%m-%d')
        else:
            agora = datetime.now()
            data_base = datetime(agora.year, agora.month, 1)

        data_inicio = datetime(data_base.year, data_base.month, 1, 0, 0, 0, 0)
        if data_base.month == 12:
            prox = datetime(data_base.year + 1, 1, 1)
        else:
            prox = datetime(data_base.year, data_base.month + 1, 1)
        data_fim = prox - timedelta(microseconds=1)

        query_base, profissional_expr, _ = _base_producao_query(data_inicio, data_fim)
        ranking_rows = (
            query_base
            .with_entities(
                profissional_expr.label('profissional'),
                func.count(ItemServico.id).label('quantidade_servicos'),
                func.coalesce(func.sum(ItemServico.valor_servico), 0.0).label('valor_total'),
                func.coalesce(func.avg(ItemServico.valor_servico), 0.0).label('media_por_servico')
            )
            .group_by(profissional_expr)
            .order_by(func.coalesce(func.sum(ItemServico.valor_servico), 0.0).desc())
            .all()
        )

        pagamentos_rows = (
            Ordem.query
            .with_entities(
                func.coalesce(func.nullif(func.trim(Ordem.forma_pagamento), ''), 'Não informado').label('forma_pagamento'),
                func.coalesce(func.sum(Ordem.total_geral), 0.0).label('valor_total'),
                func.count(Ordem.id).label('quantidade')
            )
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .group_by('forma_pagamento')
            .order_by(func.coalesce(func.sum(Ordem.total_geral), 0.0).desc())
            .all()
        )

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Periodo', f'{data_inicio.strftime("%Y-%m-%d")} a {data_fim.strftime("%Y-%m-%d")}'])
        writer.writerow([])
        writer.writerow(['Producao por profissional'])
        writer.writerow(['Profissional', 'Quantidade de servicos', 'Valor total', 'Media por servico'])
        for row in ranking_rows:
            writer.writerow([
                row.profissional,
                int(row.quantidade_servicos or 0),
                f'{float(row.valor_total or 0):.2f}',
                f'{float(row.media_por_servico or 0):.2f}'
            ])

        writer.writerow([])
        writer.writerow(['Resumo por forma de pagamento'])
        writer.writerow(['Forma de pagamento', 'Quantidade', 'Valor total'])
        for row in pagamentos_rows:
            writer.writerow([
                row.forma_pagamento,
                int(row.quantidade or 0),
                f'{float(row.valor_total or 0):.2f}'
            ])

        csv_content = buffer.getvalue()
        buffer.close()
        nome_arquivo = f'producao_pagamentos_{data_inicio.strftime("%Y_%m")}.csv'
        return Response(
            csv_content,
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={nome_arquivo}'}
        )
    except ValueError:
        return jsonify({'erro': 'Formato inválido. Use mes=YYYY-MM.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/exportar-excel', methods=['GET'])
def exportar_excel_producao_pagamentos():
    """
    Exporta Excel mensal para contador com:
    - produção por profissional (aba 1)
    - resumo por forma de pagamento (aba 2)
    """
    try:
        import pandas as pd
        from models import ItemServico, Ordem, Saida, Cliente
        engine_excel = _excel_engine_disponivel()
        if not engine_excel:
            return jsonify({'erro': 'Biblioteca para Excel não instalada. Instale openpyxl ou xlsxwriter.'}), 500

        mes = (request.args.get('mes') or '').strip()  # YYYY-MM
        if mes:
            data_base = datetime.strptime(f'{mes}-01', '%Y-%m-%d')
        else:
            agora = datetime.now()
            data_base = datetime(agora.year, agora.month, 1)

        data_inicio = datetime(data_base.year, data_base.month, 1, 0, 0, 0, 0)
        if data_base.month == 12:
            prox = datetime(data_base.year + 1, 1, 1)
        else:
            prox = datetime(data_base.year, data_base.month + 1, 1)
        data_fim = prox - timedelta(microseconds=1)

        query_base, profissional_expr, _ = _base_producao_query(data_inicio, data_fim)
        ranking_rows = (
            query_base
            .with_entities(
                profissional_expr.label('profissional'),
                func.count(ItemServico.id).label('quantidade_servicos'),
                func.coalesce(func.sum(ItemServico.valor_servico), 0.0).label('valor_total'),
                func.coalesce(func.avg(ItemServico.valor_servico), 0.0).label('media_por_servico')
            )
            .group_by(profissional_expr)
            .order_by(func.coalesce(func.sum(ItemServico.valor_servico), 0.0).desc())
            .all()
        )

        pagamentos_rows = (
            Ordem.query
            .with_entities(
                func.coalesce(func.nullif(func.trim(Ordem.forma_pagamento), ''), 'Não informado').label('forma_pagamento'),
                func.coalesce(func.sum(Ordem.total_geral), 0.0).label('valor_total'),
                func.count(Ordem.id).label('quantidade')
            )
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .group_by('forma_pagamento')
            .order_by(func.coalesce(func.sum(Ordem.total_geral), 0.0).desc())
            .all()
        )

        saidas_rows = (
            Saida.query
            .filter(Saida.data >= data_inicio.date(), Saida.data <= data_fim.date())
            .order_by(Saida.data.asc())
            .all()
        )

        os_rows = (
            Ordem.query
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .order_by(Ordem.data_conclusao.asc())
            .all()
        )

        faturamento_bruto = float(sum(float(r.valor_total or 0) for r in pagamentos_rows))
        total_saidas = float(sum(float(s.valor or 0) for s in saidas_rows))
        quantidade_os = len(os_rows)
        ticket_medio = (faturamento_bruto / quantidade_os) if quantidade_os else 0.0

        df_producao = pd.DataFrame([{
            'Profissional': row.profissional,
            'Quantidade de servicos': int(row.quantidade_servicos or 0),
            'Valor total': float(row.valor_total or 0),
            'Media por servico': float(row.media_por_servico or 0)
        } for row in ranking_rows])

        df_pagamentos = pd.DataFrame([{
            'Forma de pagamento': row.forma_pagamento,
            'Quantidade': int(row.quantidade or 0),
            'Valor total': float(row.valor_total or 0)
        } for row in pagamentos_rows])

        dados_os = []
        for o in os_rows:
            cliente_nome = '---'
            placa = '---'
            if o.cliente_id:
                cli = Cliente.query.get(o.cliente_id)
                if cli:
                    cliente_nome = cli.nome_cliente or '---'
                    placa = cli.placa or '---'
            dados_os.append({
                'OS': o.id,
                'Data conclusao': o.data_conclusao,
                'Cliente': cliente_nome,
                'Placa': placa,
                'Profissional': (o.profissional_responsavel or 'Nao informado'),
                'Forma pagamento': (o.forma_pagamento or 'Não informado'),
                'Total servicos': float(o.total_servicos or 0),
                'Total pecas': float(o.total_pecas or 0),
                'Total geral': float(o.total_geral or 0)
            })
        df_os = pd.DataFrame(dados_os)

        df_saidas = pd.DataFrame([{
            'Data': s.data,
            'Categoria': s.categoria or 'Outros',
            'Descricao': s.descricao or '',
            'Valor': float(s.valor or 0)
        } for s in saidas_rows])

        df_resumo = pd.DataFrame([
            {'Campo': 'Empresa', 'Valor': 'Oficina 39'},
            {'Campo': 'Periodo inicio', 'Valor': data_inicio.strftime('%Y-%m-%d')},
            {'Campo': 'Periodo fim', 'Valor': data_fim.strftime('%Y-%m-%d')},
            {'Campo': 'Emitido em', 'Valor': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'Campo': 'Faturamento bruto', 'Valor': faturamento_bruto},
            {'Campo': 'Total de saidas', 'Valor': total_saidas},
            {'Campo': 'Saldo operacional', 'Valor': faturamento_bruto - total_saidas},
            {'Campo': 'Quantidade OS concluidas', 'Valor': quantidade_os},
            {'Campo': 'Ticket medio', 'Valor': ticket_medio}
        ])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine=engine_excel) as writer:
            # Cabeçalho/Resumo para contador
            df_resumo.to_excel(writer, sheet_name='Resumo_Contabil', index=False)

            if df_producao.empty:
                df_producao = pd.DataFrame([{
                    'Profissional': 'Sem dados',
                    'Quantidade de servicos': 0,
                    'Valor total': 0.0,
                    'Media por servico': 0.0
                }])
            df_producao.to_excel(writer, sheet_name='Producao', index=False)

            if df_pagamentos.empty:
                df_pagamentos = pd.DataFrame([{
                    'Forma de pagamento': 'Sem dados',
                    'Quantidade': 0,
                    'Valor total': 0.0
                }])
            df_pagamentos.to_excel(writer, sheet_name='Pagamentos', index=False)

            if df_os.empty:
                df_os = pd.DataFrame([{
                    'OS': '',
                    'Data conclusao': '',
                    'Cliente': 'Sem dados',
                    'Placa': '',
                    'Profissional': '',
                    'Forma pagamento': '',
                    'Total servicos': 0.0,
                    'Total pecas': 0.0,
                    'Total geral': 0.0
                }])
            df_os.to_excel(writer, sheet_name='OS_Concluidas', index=False)

            if df_saidas.empty:
                df_saidas = pd.DataFrame([{
                    'Data': '',
                    'Categoria': 'Sem dados',
                    'Descricao': '',
                    'Valor': 0.0
                }])
            df_saidas.to_excel(writer, sheet_name='Saidas', index=False)

            _aplicar_estilo_excel(writer, engine_excel, [
                {
                    'sheet_name': 'Resumo_Contabil',
                    'df': df_resumo,
                    'currency_cols': ['Valor']
                },
                {
                    'sheet_name': 'Producao',
                    'df': df_producao,
                    'currency_cols': ['Valor total', 'Media por servico'],
                    'integer_cols': ['Quantidade de servicos']
                },
                {
                    'sheet_name': 'Pagamentos',
                    'df': df_pagamentos,
                    'currency_cols': ['Valor total'],
                    'integer_cols': ['Quantidade']
                },
                {
                    'sheet_name': 'OS_Concluidas',
                    'df': df_os,
                    'currency_cols': ['Total servicos', 'Total pecas', 'Total geral'],
                    'datetime_cols': ['Data conclusao']
                },
                {
                    'sheet_name': 'Saidas',
                    'df': df_saidas,
                    'currency_cols': ['Valor'],
                    'date_cols': ['Data']
                }
            ])

        output.seek(0)
        nome_arquivo = f'producao_pagamentos_{data_inicio.strftime("%Y_%m")}.xlsx'
        logger.info("Exportacao Excel mensal gerada: %s", nome_arquivo)
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={nome_arquivo}'}
        )
    except ValueError:
        return jsonify({'erro': 'Formato inválido. Use mes=YYYY-MM.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/producao-profissionais/exportar-excel-profissional', methods=['GET'])
def exportar_excel_profissional_periodo():
    """
    Exporta Excel de um profissional em período customizado (máx. 31 dias).
    """
    try:
        import pandas as pd

        engine_excel = _excel_engine_disponivel()
        if not engine_excel:
            return jsonify({'erro': 'Biblioteca para Excel não instalada. Instale openpyxl ou xlsxwriter.'}), 500

        profissional = (request.args.get('profissional') or '').strip()
        data_inicio_str = (request.args.get('data_inicio') or '').strip()
        data_fim_str = (request.args.get('data_fim') or '').strip()

        if not profissional or not data_inicio_str or not data_fim_str:
            return jsonify({'erro': 'Informe profissional, data_inicio e data_fim.'}), 400
        if relatorio_repository.total_profissionais_ativos() == 0:
            return jsonify({'erro': 'Não há profissionais cadastrados/ativos para gerar relatório.'}), 400

        data_inicio = _parse_data(data_inicio_str)
        data_fim = _parse_data(data_fim_str, fim_do_dia=True)
        if not data_inicio or not data_fim:
            return jsonify({'erro': 'Datas inválidas.'}), 400
        if data_fim < data_inicio:
            return jsonify({'erro': 'data_fim deve ser maior ou igual à data_inicio.'}), 400
        dias_intervalo = (data_fim.date() - data_inicio.date()).days + 1
        if dias_intervalo > 31:
            return jsonify({'erro': 'O período máximo permitido é de 31 dias.'}), 400

        profissional_resolvido = _resolver_profissional_ativo(profissional)
        if not profissional_resolvido:
            return jsonify({'erro': 'Profissional não encontrado na lista de cadastrados/ativos.'}), 404

        nome_resolvido = profissional_resolvido.nome
        bloco = _detalhes_resumo_profissional(nome_resolvido, data_inicio, data_fim, limite=10000)
        resumo = bloco.get('resumo', {})
        servicos = bloco.get('servicos', [])

        df_resumo = pd.DataFrame([
            {'Campo': 'Empresa', 'Valor': 'Oficina 39'},
            {'Campo': 'Profissional', 'Valor': nome_resolvido},
            {'Campo': 'Período início', 'Valor': data_inicio.strftime('%Y-%m-%d')},
            {'Campo': 'Período fim', 'Valor': data_fim.strftime('%Y-%m-%d')},
            {'Campo': 'Emitido em', 'Valor': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'Campo': 'Quantidade de serviços', 'Valor': int(resumo.get('quantidade_servicos', 0) or 0)},
            {'Campo': 'Valor total', 'Valor': float(resumo.get('valor_total', 0) or 0)},
            {'Campo': 'Média por serviço', 'Valor': float(resumo.get('media_por_servico', 0) or 0)}
        ])

        df_servicos = pd.DataFrame([{
            'OS': s.get('ordem_id'),
            'Data': s.get('data_referencia'),
            'Cliente': s.get('cliente'),
            'Servico': s.get('descricao_servico'),
            'Valor': float(s.get('valor_servico', 0) or 0)
        } for s in servicos])
        if not df_servicos.empty:
            df_servicos['Data'] = pd.to_datetime(df_servicos['Data'], format='%d/%m/%Y', errors='coerce')
        if df_servicos.empty:
            df_servicos = pd.DataFrame([{
                'OS': '',
                'Data': '',
                'Cliente': 'Sem dados',
                'Servico': '',
                'Valor': 0.0
            }])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine=engine_excel) as writer:
            df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
            df_servicos.to_excel(writer, sheet_name='Servicos', index=False)

            _aplicar_estilo_excel(writer, engine_excel, [
                {
                    'sheet_name': 'Resumo',
                    'df': df_resumo,
                    'currency_cols': ['Valor']
                },
                {
                    'sheet_name': 'Servicos',
                    'df': df_servicos,
                    'currency_cols': ['Valor'],
                    'date_cols': ['Data']
                }
            ])

        output.seek(0)
        nome_arquivo = f'producao_profissional_{nome_resolvido.replace(" ", "_")}_{data_fim.strftime("%Y_%m_%d")}.xlsx'
        logger.info(
            "Exportacao Excel profissional gerada: %s (profissional=%s, inicio=%s, fim=%s)",
            nome_arquivo,
            nome_resolvido,
            data_inicio.strftime('%Y-%m-%d'),
            data_fim.strftime('%Y-%m-%d')
        )
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={nome_arquivo}'}
        )
    except ValueError:
        return jsonify({'erro': 'Formato inválido. Use datas em YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/contabilidade-geral', methods=['GET'])
def relatorio_contabilidade_geral():
    """
    Resumo contábil mensal para apoio ao contador.
    """
    try:
        from models import Ordem, Saida

        mes = (request.args.get('mes') or '').strip()  # YYYY-MM
        if mes:
            data_base = datetime.strptime(f'{mes}-01', '%Y-%m-%d')
        else:
            agora = datetime.now()
            data_base = datetime(agora.year, agora.month, 1)

        data_inicio = datetime(data_base.year, data_base.month, 1, 0, 0, 0, 0)
        if data_base.month == 12:
            prox = datetime(data_base.year + 1, 1, 1)
        else:
            prox = datetime(data_base.year, data_base.month + 1, 1)
        data_fim = prox - timedelta(microseconds=1)

        # Faturamento (ordens concluídas/garantia no mês)
        faturamento_row = (
            Ordem.query
            .with_entities(func.coalesce(func.sum(Ordem.total_geral), 0.0))
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .first()
        )
        faturamento_bruto = float((faturamento_row[0] if faturamento_row else 0) or 0)

        quantidade_os = (
            Ordem.query
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .count()
        )

        # Saídas financeiras no mês
        saidas_row = (
            Saida.query
            .with_entities(func.coalesce(func.sum(Saida.valor), 0.0))
            .filter(Saida.data >= data_inicio.date(), Saida.data <= data_fim.date())
            .first()
        )
        total_saidas = float((saidas_row[0] if saidas_row else 0) or 0)

        pagamentos_rows = (
            Ordem.query
            .with_entities(
                func.coalesce(func.nullif(func.trim(Ordem.forma_pagamento), ''), 'Não informado').label('forma_pagamento'),
                func.coalesce(func.sum(Ordem.total_geral), 0.0).label('valor_total'),
                func.count(Ordem.id).label('quantidade')
            )
            .filter(Ordem.status.in_(['Concluído', 'Garantia']))
            .filter(Ordem.data_conclusao >= data_inicio, Ordem.data_conclusao <= data_fim)
            .group_by('forma_pagamento')
            .order_by(func.coalesce(func.sum(Ordem.total_geral), 0.0).desc())
            .all()
        )

        ticket_medio = (faturamento_bruto / quantidade_os) if quantidade_os else 0.0

        return jsonify({
            'mes_referencia': data_inicio.strftime('%Y-%m'),
            'periodo': {
                'inicio': data_inicio.strftime('%Y-%m-%d'),
                'fim': data_fim.strftime('%Y-%m-%d')
            },
            'faturamento_bruto': faturamento_bruto,
            'total_saidas': total_saidas,
            'saldo_operacional': faturamento_bruto - total_saidas,
            'quantidade_os': int(quantidade_os or 0),
            'ticket_medio': ticket_medio,
            'pagamentos': [{
                'forma_pagamento': row.forma_pagamento,
                'valor_total': float(row.valor_total or 0),
                'quantidade': int(row.quantidade or 0)
            } for row in pagamentos_rows]
        })
    except ValueError:
        return jsonify({'erro': 'Formato inválido. Use mes=YYYY-MM.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/operacional-servicos-pecas-saidas', methods=['GET'])
def relatorio_operacional_servicos_pecas_saidas():
    """
    Relatório operacional enxuto para controle diário:
    - serviços aplicados
    - peças lançadas
    - saídas de caixa
    """
    try:
        from models import Ordem, ItemServico, ItemPeca, Saida

        data_inicio = _parse_data(request.args.get('data_inicio'))
        data_fim = _parse_data(request.args.get('data_fim'), fim_do_dia=True)
        if not data_inicio or not data_fim:
            hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            data_inicio = hoje
            data_fim = hoje.replace(hour=23, minute=59, second=59, microsecond=999999)
        if data_fim < data_inicio:
            return jsonify({'erro': 'data_fim deve ser maior ou igual à data_inicio.'}), 400

        # Agregação de serviços no período.
        servicos_rows = (
            ItemServico.query
            .join(Ordem, ItemServico.ordem_id == Ordem.id)
            .with_entities(
                Ordem.id.label('ordem_id'),
                func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).label('data_ref'),
                func.coalesce(func.nullif(func.trim(ItemServico.nome_profissional), ''), func.nullif(func.trim(Ordem.profissional_responsavel), ''), 'Nao informado').label('profissional'),
                ItemServico.descricao_servico.label('descricao'),
                func.coalesce(ItemServico.valor_servico, 0.0).label('valor')
            )
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) >= data_inicio)
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) <= data_fim)
            .order_by(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).desc())
            .all()
        )

        # Agregação de peças no período.
        pecas_rows = (
            ItemPeca.query
            .join(Ordem, ItemPeca.ordem_id == Ordem.id)
            .with_entities(
                Ordem.id.label('ordem_id'),
                func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).label('data_ref'),
                ItemPeca.descricao_peca.label('descricao'),
                func.coalesce(ItemPeca.quantidade, 0.0).label('quantidade'),
                func.coalesce(ItemPeca.valor_unitario, 0.0).label('valor_unitario')
            )
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) >= data_inicio)
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) <= data_fim)
            .order_by(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).desc())
            .all()
        )

        # Saídas financeiras já lançadas no período.
        saidas_rows = (
            Saida.query
            .with_entities(
                Saida.data.label('data_ref'),
                func.coalesce(Saida.categoria, 'Outros').label('categoria'),
                func.coalesce(Saida.descricao, '').label('descricao'),
                func.coalesce(Saida.valor, 0.0).label('valor')
            )
            .filter(Saida.data >= data_inicio, Saida.data <= data_fim)
            .order_by(Saida.data.desc())
            .all()
        )

        total_servicos = float(sum(float(r.valor or 0) for r in servicos_rows))
        total_pecas = float(sum(float((r.quantidade or 0) * (r.valor_unitario or 0)) for r in pecas_rows))
        total_saidas = float(sum(float(r.valor or 0) for r in saidas_rows))

        return jsonify({
            'periodo': {
                'inicio': data_inicio.strftime('%Y-%m-%d'),
                'fim': data_fim.strftime('%Y-%m-%d')
            },
            'resumo': {
                'quantidade_servicos': len(servicos_rows),
                'valor_servicos': total_servicos,
                'quantidade_pecas': len(pecas_rows),
                'valor_pecas': total_pecas,
                'quantidade_saidas': len(saidas_rows),
                'valor_saidas': total_saidas
            },
            'servicos': [{
                'ordem_id': r.ordem_id,
                'data_referencia': (r.data_ref.strftime('%d/%m/%Y') if r.data_ref else '---'),
                'profissional': r.profissional or 'Nao informado',
                'descricao': r.descricao or '---',
                'valor': float(r.valor or 0)
            } for r in servicos_rows],
            'pecas': [{
                'ordem_id': r.ordem_id,
                'data_referencia': (r.data_ref.strftime('%d/%m/%Y') if r.data_ref else '---'),
                'descricao': r.descricao or '---',
                'quantidade': float(r.quantidade or 0),
                'valor_total': float((r.quantidade or 0) * (r.valor_unitario or 0))
            } for r in pecas_rows],
            'saidas': [{
                'data_referencia': (r.data_ref.strftime('%d/%m/%Y') if r.data_ref else '---'),
                'categoria': r.categoria or 'Outros',
                'descricao': r.descricao or '---',
                'valor': float(r.valor or 0)
            } for r in saidas_rows]
        })
    except ValueError:
        return jsonify({'erro': 'Formato inválido. Use datas em YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@relatorios_bp.route('/operacional-servicos-pecas-saidas/exportar-excel', methods=['GET'])
def exportar_excel_operacional_servicos_pecas_saidas():
    """Exporta planilha operacional com abas: resumo, serviços, peças e saídas."""
    try:
        import pandas as pd
        from models import Ordem, ItemServico, ItemPeca, Saida

        engine_excel = _excel_engine_disponivel()
        if not engine_excel:
            return jsonify({'erro': 'Biblioteca para Excel não instalada. Instale openpyxl ou xlsxwriter.'}), 500

        data_inicio = _parse_data(request.args.get('data_inicio'))
        data_fim = _parse_data(request.args.get('data_fim'), fim_do_dia=True)
        if not data_inicio or not data_fim:
            hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            data_inicio = hoje
            data_fim = hoje.replace(hour=23, minute=59, second=59, microsecond=999999)
        if data_fim < data_inicio:
            return jsonify({'erro': 'data_fim deve ser maior ou igual à data_inicio.'}), 400

        servicos_rows = (
            ItemServico.query
            .join(Ordem, ItemServico.ordem_id == Ordem.id)
            .with_entities(
                Ordem.id.label('ordem_id'),
                func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).label('data_ref'),
                func.coalesce(func.nullif(func.trim(ItemServico.nome_profissional), ''), func.nullif(func.trim(Ordem.profissional_responsavel), ''), 'Nao informado').label('profissional'),
                ItemServico.descricao_servico.label('descricao'),
                func.coalesce(ItemServico.valor_servico, 0.0).label('valor')
            )
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) >= data_inicio)
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) <= data_fim)
            .order_by(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).desc())
            .all()
        )
        pecas_rows = (
            ItemPeca.query
            .join(Ordem, ItemPeca.ordem_id == Ordem.id)
            .with_entities(
                Ordem.id.label('ordem_id'),
                func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).label('data_ref'),
                ItemPeca.descricao_peca.label('descricao'),
                func.coalesce(ItemPeca.quantidade, 0.0).label('quantidade'),
                func.coalesce(ItemPeca.valor_unitario, 0.0).label('valor_unitario')
            )
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) >= data_inicio)
            .filter(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada) <= data_fim)
            .order_by(func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada).desc())
            .all()
        )
        saidas_rows = (
            Saida.query
            .with_entities(
                Saida.data.label('data_ref'),
                func.coalesce(Saida.categoria, 'Outros').label('categoria'),
                func.coalesce(Saida.descricao, '').label('descricao'),
                func.coalesce(Saida.valor, 0.0).label('valor')
            )
            .filter(Saida.data >= data_inicio, Saida.data <= data_fim)
            .order_by(Saida.data.desc())
            .all()
        )

        total_servicos = float(sum(float(r.valor or 0) for r in servicos_rows))
        total_pecas = float(sum(float((r.quantidade or 0) * (r.valor_unitario or 0)) for r in pecas_rows))
        total_saidas = float(sum(float(r.valor or 0) for r in saidas_rows))

        df_resumo = pd.DataFrame([
            {'Campo': 'Relatorio', 'Valor': 'Operacional - Servicos, Pecas e Saidas'},
            {'Campo': 'Periodo inicio', 'Valor': data_inicio.strftime('%Y-%m-%d')},
            {'Campo': 'Periodo fim', 'Valor': data_fim.strftime('%Y-%m-%d')},
            {'Campo': 'Emitido em', 'Valor': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'Campo': 'Qtd servicos', 'Valor': len(servicos_rows)},
            {'Campo': 'Total servicos', 'Valor': total_servicos},
            {'Campo': 'Qtd pecas', 'Valor': len(pecas_rows)},
            {'Campo': 'Total pecas', 'Valor': total_pecas},
            {'Campo': 'Qtd saidas', 'Valor': len(saidas_rows)},
            {'Campo': 'Total saidas', 'Valor': total_saidas}
        ])

        df_servicos = pd.DataFrame([{
            'OS': r.ordem_id,
            'Data': r.data_ref,
            'Profissional': r.profissional or 'Nao informado',
            'Servico': r.descricao or '---',
            'Valor': float(r.valor or 0)
        } for r in servicos_rows])
        if df_servicos.empty:
            df_servicos = pd.DataFrame([{'OS': '', 'Data': '', 'Profissional': 'Sem dados', 'Servico': '', 'Valor': 0.0}])

        df_pecas = pd.DataFrame([{
            'OS': r.ordem_id,
            'Data': r.data_ref,
            'Peca': r.descricao or '---',
            'Quantidade': float(r.quantidade or 0),
            'Valor total': float((r.quantidade or 0) * (r.valor_unitario or 0))
        } for r in pecas_rows])
        if df_pecas.empty:
            df_pecas = pd.DataFrame([{'OS': '', 'Data': '', 'Peca': 'Sem dados', 'Quantidade': 0.0, 'Valor total': 0.0}])

        df_saidas = pd.DataFrame([{
            'Data': r.data_ref,
            'Categoria': r.categoria or 'Outros',
            'Descricao': r.descricao or '',
            'Valor': float(r.valor or 0)
        } for r in saidas_rows])
        if df_saidas.empty:
            df_saidas = pd.DataFrame([{'Data': '', 'Categoria': 'Sem dados', 'Descricao': '', 'Valor': 0.0}])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine=engine_excel) as writer:
            df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
            df_servicos.to_excel(writer, sheet_name='Servicos', index=False)
            df_pecas.to_excel(writer, sheet_name='Pecas', index=False)
            df_saidas.to_excel(writer, sheet_name='Saidas', index=False)

            _aplicar_estilo_excel(writer, engine_excel, [
                {'sheet_name': 'Resumo', 'df': df_resumo, 'currency_cols': ['Valor']},
                {'sheet_name': 'Servicos', 'df': df_servicos, 'currency_cols': ['Valor'], 'date_cols': ['Data']},
                {'sheet_name': 'Pecas', 'df': df_pecas, 'currency_cols': ['Valor total'], 'date_cols': ['Data']},
                {'sheet_name': 'Saidas', 'df': df_saidas, 'currency_cols': ['Valor'], 'date_cols': ['Data']}
            ])

        output.seek(0)
        nome_arquivo = f'operacional_servicos_pecas_saidas_{data_inicio.strftime("%Y_%m_%d")}_{data_fim.strftime("%Y_%m_%d")}.xlsx'
        logger.info(
            "Exportacao Excel operacional gerada: %s (inicio=%s, fim=%s)",
            nome_arquivo,
            data_inicio.strftime('%Y-%m-%d'),
            data_fim.strftime('%Y-%m-%d')
        )
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={nome_arquivo}'}
        )
    except ValueError:
        return jsonify({'erro': 'Formato inválido. Use datas em YYYY-MM-DD.'}), 400
    except Exception as e:
        return jsonify({'erro': str(e)}), 500
