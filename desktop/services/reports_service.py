from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from models import ItemPeca, ItemServico, Ordem, Saida
from repositories import relatorio_repository


HEADER_FILL = PatternFill(fill_type="solid", fgColor="0F4C5C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
BODY_ALIGN = Alignment(vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _to_datetime_range(start_date: date, end_date: date) -> tuple[datetime, datetime]:
    start = datetime.combine(start_date, datetime.min.time())
    end = datetime.combine(end_date, datetime.max.time())
    return start, end


def _month_period(reference_date: date) -> tuple[datetime, datetime]:
    start = datetime(reference_date.year, reference_date.month, 1, 0, 0, 0, 0)
    if reference_date.month == 12:
        next_month = datetime(reference_date.year + 1, 1, 1)
    else:
        next_month = datetime(reference_date.year, reference_date.month + 1, 1)
    return start, next_month - timedelta(microseconds=1)


def list_active_professionals() -> list[str]:
    rows = relatorio_repository.listar_profissionais_ativos("", limite=200)
    return [row.nome for row in rows if (row.nome or "").strip()]


def get_production_summary(professional_name: str, start_date: date, end_date: date) -> dict:
    if not professional_name:
        raise ValueError("Selecione um profissional cadastrado.")
    start, end = _to_datetime_range(start_date, end_date)
    days = (end.date() - start.date()).days + 1
    if days > 31:
        raise ValueError("O período máximo permitido é de 31 dias.")

    professional = relatorio_repository.resolver_profissional_ativo(professional_name)
    if not professional:
        raise ValueError("Profissional não encontrado na lista de ativos.")

    details = relatorio_repository.detalhes_resumo_profissional(professional.nome, start, end, limite=10000)
    return {
        "profissional": professional.nome,
        "periodo": {"inicio": start.date().isoformat(), "fim": end.date().isoformat(), "dias": days},
        "resumo": details.get("resumo", {}),
        "servicos": details.get("servicos", []),
    }


def get_accounting_summary(reference_date: date) -> dict:
    start, end = _month_period(reference_date)

    faturamento_row = (
        Ordem.query.with_entities(func.coalesce(func.sum(Ordem.total_geral), 0.0))
        .filter(Ordem.status.in_(["Concluído", "Garantia"]))
        .filter(Ordem.data_conclusao >= start, Ordem.data_conclusao <= end)
        .first()
    )
    gross_revenue = float((faturamento_row[0] if faturamento_row else 0) or 0)

    total_orders = (
        Ordem.query.filter(Ordem.status.in_(["Concluído", "Garantia"]))
        .filter(Ordem.data_conclusao >= start, Ordem.data_conclusao <= end)
        .count()
    )

    outputs_row = (
        Saida.query.with_entities(func.coalesce(func.sum(Saida.valor), 0.0))
        .filter(Saida.data >= start.date(), Saida.data <= end.date())
        .first()
    )
    total_outputs = float((outputs_row[0] if outputs_row else 0) or 0)

    payments_rows = (
        Ordem.query.with_entities(
            func.coalesce(func.nullif(func.trim(Ordem.forma_pagamento), ""), "Não informado").label("forma_pagamento"),
            func.coalesce(func.sum(Ordem.total_geral), 0.0).label("valor_total"),
            func.count(Ordem.id).label("quantidade"),
        )
        .filter(Ordem.status.in_(["Concluído", "Garantia"]))
        .filter(Ordem.data_conclusao >= start, Ordem.data_conclusao <= end)
        .group_by("forma_pagamento")
        .order_by(func.coalesce(func.sum(Ordem.total_geral), 0.0).desc())
        .all()
    )

    average_ticket = gross_revenue / total_orders if total_orders else 0.0
    return {
        "mes_referencia": start.strftime("%Y-%m"),
        "periodo": {"inicio": start.date().isoformat(), "fim": end.date().isoformat()},
        "faturamento_bruto": gross_revenue,
        "total_saidas": total_outputs,
        "saldo_operacional": gross_revenue - total_outputs,
        "quantidade_os": int(total_orders or 0),
        "ticket_medio": average_ticket,
        "pagamentos": [
            {
                "forma_pagamento": row.forma_pagamento,
                "valor_total": float(row.valor_total or 0),
                "quantidade": int(row.quantidade or 0),
            }
            for row in payments_rows
        ],
    }


def get_operational_summary(start_date: date, end_date: date) -> dict:
    start, end = _to_datetime_range(start_date, end_date)
    if end < start:
        raise ValueError("Data fim deve ser maior ou igual à data início.")

    reference_expr = func.coalesce(Ordem.data_conclusao, Ordem.data_emissao, Ordem.data_entrada)

    services_rows = (
        ItemServico.query.join(Ordem, ItemServico.ordem_id == Ordem.id)
        .with_entities(
            Ordem.id.label("ordem_id"),
            reference_expr.label("data_ref"),
            func.coalesce(
                func.nullif(func.trim(ItemServico.nome_profissional), ""),
                func.nullif(func.trim(Ordem.profissional_responsavel), ""),
                "Não informado",
            ).label("profissional"),
            ItemServico.descricao_servico.label("descricao"),
            func.coalesce(ItemServico.valor_servico, 0.0).label("valor"),
        )
        .filter(reference_expr >= start)
        .filter(reference_expr <= end)
        .order_by(reference_expr.desc())
        .all()
    )

    parts_rows = (
        ItemPeca.query.join(Ordem, ItemPeca.ordem_id == Ordem.id)
        .with_entities(
            Ordem.id.label("ordem_id"),
            reference_expr.label("data_ref"),
            ItemPeca.descricao_peca.label("descricao"),
            func.coalesce(ItemPeca.quantidade, 0.0).label("quantidade"),
            func.coalesce(ItemPeca.valor_unitario, 0.0).label("valor_unitario"),
        )
        .filter(reference_expr >= start)
        .filter(reference_expr <= end)
        .order_by(reference_expr.desc())
        .all()
    )

    outputs_rows = (
        Saida.query.with_entities(
            Saida.data.label("data_ref"),
            func.coalesce(Saida.categoria, "Outros").label("categoria"),
            func.coalesce(Saida.descricao, "").label("descricao"),
            func.coalesce(Saida.valor, 0.0).label("valor"),
        )
        .filter(Saida.data >= start.date(), Saida.data <= end.date())
        .order_by(Saida.data.desc())
        .all()
    )

    total_services = float(sum(float(row.valor or 0) for row in services_rows))
    total_parts = float(sum(float((row.quantidade or 0) * (row.valor_unitario or 0)) for row in parts_rows))
    total_outputs = float(sum(float(row.valor or 0) for row in outputs_rows))

    return {
        "periodo": {"inicio": start.date().isoformat(), "fim": end.date().isoformat()},
        "resumo": {
            "quantidade_servicos": len(services_rows),
            "valor_servicos": total_services,
            "quantidade_pecas": len(parts_rows),
            "valor_pecas": total_parts,
            "quantidade_saidas": len(outputs_rows),
            "valor_saidas": total_outputs,
        },
        "servicos": [
            {
                "ordem_id": row.ordem_id,
                "data_referencia": row.data_ref.strftime("%d/%m/%Y") if row.data_ref else "---",
                "profissional": row.profissional or "Não informado",
                "descricao": row.descricao or "---",
                "valor": float(row.valor or 0),
            }
            for row in services_rows
        ],
        "pecas": [
            {
                "ordem_id": row.ordem_id,
                "data_referencia": row.data_ref.strftime("%d/%m/%Y") if row.data_ref else "---",
                "descricao": row.descricao or "---",
                "quantidade": float(row.quantidade or 0),
                "valor_total": float((row.quantidade or 0) * (row.valor_unitario or 0)),
            }
            for row in parts_rows
        ],
        "saidas": [
            {
                "data_referencia": row.data_ref.strftime("%d/%m/%Y") if row.data_ref else "---",
                "categoria": row.categoria or "Outros",
                "descricao": row.descricao or "---",
                "valor": float(row.valor or 0),
            }
            for row in outputs_rows
        ],
    }


def default_filename(report_kind: str, **kwargs) -> str:
    if report_kind == "producao":
        professional = kwargs["professional"].replace(" ", "_")
        end_date = kwargs["end_date"].strftime("%Y_%m_%d")
        return f"producao_profissional_{professional}_{end_date}.xlsx"
    if report_kind == "contabilidade":
        month = kwargs["month"].strftime("%Y_%m")
        return f"contabilidade_geral_{month}.xlsx"
    if report_kind == "operacional":
        start = kwargs["start_date"].strftime("%Y_%m_%d")
        end = kwargs["end_date"].strftime("%Y_%m_%d")
        return f"operacional_servicos_pecas_saidas_{start}_{end}.xlsx"
    raise ValueError("Tipo de relatório inválido.")


def _write_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list], currency_cols: set[str] | None = None) -> None:
    sheet = workbook.create_sheet(title=title)
    currency_cols = currency_cols or set()
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_index = {name: idx + 1 for idx, name in enumerate(headers)}
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = BODY_ALIGN
        for column_name in currency_cols:
            col_idx = header_index.get(column_name)
            if col_idx:
                row[col_idx - 1].number_format = 'R$ #,##0.00'
                row[col_idx - 1].alignment = CENTER_ALIGN

    for col_idx, column_name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(column_name))
        for row_idx in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        sheet.column_dimensions[letter].width = min(max(max_len, 10) + 2, 48)


def _create_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def _save_workbook(workbook: Workbook, output_path: str | Path) -> Path:
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


def export_production_excel(professional_name: str, start_date: date, end_date: date, output_path: str | Path) -> Path:
    report = get_production_summary(professional_name, start_date, end_date)
    summary = report["resumo"]
    workbook = _create_workbook()

    _write_sheet(
        workbook,
        "Resumo",
        ["Campo", "Valor"],
        [
            ["Empresa", "Oficina 39"],
            ["Profissional", report["profissional"]],
            ["Período início", report["periodo"]["inicio"]],
            ["Período fim", report["periodo"]["fim"]],
            ["Quantidade de serviços", int(summary.get("quantidade_servicos", 0) or 0)],
            ["Valor total", float(summary.get("valor_total", 0) or 0)],
            ["Média por serviço", float(summary.get("media_por_servico", 0) or 0)],
        ],
        currency_cols={"Valor"},
    )
    _write_sheet(
        workbook,
        "Servicos",
        ["OS", "Data", "Cliente", "Serviço", "Valor"],
        [
            [
                item.get("ordem_id"),
                item.get("data_referencia"),
                item.get("cliente"),
                item.get("descricao_servico"),
                float(item.get("valor_servico", 0) or 0),
            ]
            for item in report["servicos"]
        ]
        or [["", "", "Sem dados", "", 0.0]],
        currency_cols={"Valor"},
    )
    return _save_workbook(workbook, output_path)


def export_accounting_excel(reference_date: date, output_path: str | Path) -> Path:
    report = get_accounting_summary(reference_date)
    workbook = _create_workbook()

    _write_sheet(
        workbook,
        "Resumo",
        ["Campo", "Valor"],
        [
            ["Mês referência", report["mes_referencia"]],
            ["Período início", report["periodo"]["inicio"]],
            ["Período fim", report["periodo"]["fim"]],
            ["Faturamento bruto", report["faturamento_bruto"]],
            ["Total saídas", report["total_saidas"]],
            ["Saldo operacional", report["saldo_operacional"]],
            ["Quantidade OS", report["quantidade_os"]],
            ["Ticket médio", report["ticket_medio"]],
        ],
        currency_cols={"Valor"},
    )
    _write_sheet(
        workbook,
        "Pagamentos",
        ["Forma de pagamento", "Quantidade", "Valor total"],
        [
            [item["forma_pagamento"], int(item["quantidade"] or 0), float(item["valor_total"] or 0)]
            for item in report["pagamentos"]
        ]
        or [["Sem dados", 0, 0.0]],
        currency_cols={"Valor total"},
    )
    return _save_workbook(workbook, output_path)


def export_operational_excel(start_date: date, end_date: date, output_path: str | Path) -> Path:
    report = get_operational_summary(start_date, end_date)
    summary = report["resumo"]
    workbook = _create_workbook()

    _write_sheet(
        workbook,
        "Resumo",
        ["Campo", "Valor"],
        [
            ["Período início", report["periodo"]["inicio"]],
            ["Período fim", report["periodo"]["fim"]],
            ["Qtd serviços", summary["quantidade_servicos"]],
            ["Total serviços", summary["valor_servicos"]],
            ["Qtd peças", summary["quantidade_pecas"]],
            ["Total peças", summary["valor_pecas"]],
            ["Qtd saídas", summary["quantidade_saidas"]],
            ["Total saídas", summary["valor_saidas"]],
        ],
        currency_cols={"Valor"},
    )
    _write_sheet(
        workbook,
        "Servicos",
        ["OS", "Data", "Profissional", "Serviço", "Valor"],
        [
            [item["ordem_id"], item["data_referencia"], item["profissional"], item["descricao"], item["valor"]]
            for item in report["servicos"]
        ]
        or [["", "", "Sem dados", "", 0.0]],
        currency_cols={"Valor"},
    )
    _write_sheet(
        workbook,
        "Pecas",
        ["OS", "Data", "Peça", "Quantidade", "Valor total"],
        [
            [item["ordem_id"], item["data_referencia"], item["descricao"], item["quantidade"], item["valor_total"]]
            for item in report["pecas"]
        ]
        or [["", "", "Sem dados", 0.0, 0.0]],
        currency_cols={"Valor total"},
    )
    _write_sheet(
        workbook,
        "Saidas",
        ["Data", "Categoria", "Descrição", "Valor"],
        [
            [item["data_referencia"], item["categoria"], item["descricao"], item["valor"]]
            for item in report["saidas"]
        ]
        or [["", "Sem dados", "", 0.0]],
        currency_cols={"Valor"},
    )
    return _save_workbook(workbook, output_path)
