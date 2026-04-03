from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from services.order_pdf_service import generate_order_pdf_bytes, suggested_order_pdf_name


def exportar_pdf_ordem(order_id: int) -> tuple[str, bytes]:
    return suggested_order_pdf_name(order_id), generate_order_pdf_bytes(order_id)


def exportar_excel_operacional(relatorio: dict) -> tuple[str, bytes]:
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'
    _append_rows(ws_resumo, ['Campo', 'Valor'], [
        ('Relatorio', 'Operacional - Servicos, Pecas e Saidas'),
        ('Periodo inicio', relatorio['periodo']['inicio']),
        ('Periodo fim', relatorio['periodo']['fim']),
        ('Qtd servicos', relatorio['resumo']['quantidade_servicos']),
        ('Total servicos', relatorio['resumo']['valor_servicos']),
        ('Qtd pecas', relatorio['resumo']['quantidade_pecas']),
        ('Total pecas', relatorio['resumo']['valor_pecas']),
        ('Qtd saidas', relatorio['resumo']['quantidade_saidas']),
        ('Total saidas', relatorio['resumo']['valor_saidas']),
    ])

    ws_servicos = wb.create_sheet('Servicos')
    _append_rows(ws_servicos, ['OS', 'Data', 'Profissional', 'Servico', 'Valor'], [
        (
            item['ordem_id'],
            item['data_referencia'],
            item['profissional'],
            item['descricao'],
            item['valor'],
        )
        for item in relatorio['servicos']
    ] or [('', '', 'Sem dados', '', 0.0)])

    ws_pecas = wb.create_sheet('Pecas')
    _append_rows(ws_pecas, ['OS', 'Data', 'Peca', 'Quantidade', 'Valor total'], [
        (
            item['ordem_id'],
            item['data_referencia'],
            item['descricao'],
            item['quantidade'],
            item['valor_total'],
        )
        for item in relatorio['pecas']
    ] or [('', '', 'Sem dados', 0.0, 0.0)])

    ws_saidas = wb.create_sheet('Saidas')
    _append_rows(ws_saidas, ['Data', 'Categoria', 'Descricao', 'Valor'], [
        (
            item['data_referencia'],
            item['categoria'],
            item['descricao'],
            item['valor'],
        )
        for item in relatorio['saidas']
    ] or [('', 'Sem dados', '', 0.0)])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome = f"operacional_servicos_pecas_saidas_{relatorio['periodo']['inicio'].replace('-', '_')}_{relatorio['periodo']['fim'].replace('-', '_')}.xlsx"
    return nome, buffer.getvalue()


def _append_rows(ws, header, rows):
    ws.append(list(header))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))
